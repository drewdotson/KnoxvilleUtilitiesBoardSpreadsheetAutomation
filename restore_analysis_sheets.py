# The purpose of this file is to restore the pre-update format of the area metrics and month-by-month metrics 
# analysis sheets.

from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# FUNCTIONS 👇 ---------------------------------------------------------------------------------------------------#

# This function is used to return the first row of a specified sheet that does not have metric data. It takes a
# sheet name as input.
def find_row_limit(sheet):

    # Iterate over the first column column in each row.
    for cell in sheet['A']:

        # When the first row that does not have a value in the first column is found, the function will return 
        # that row's number.
        if cell.value is None:
            return(cell.row)
        else:
            continue
    
    # If there are no rows in the specified sheet without an empty first column, return the first next row's number
    return sheet.max_row + 1


# Function used for restoring each row's height and each column's width to their original values. It takes a 
# specified sheet name and its row limit as input.
def restore_height_and_width(sheet, row_limit):

    # The header row is given a height value of 58
    sheet.row_dimensions[1].height = 58

    # Each non-header row is given a height value of 22
    for row in range(2, row_limit):
        sheet.row_dimensions[row].height = 22

    # Each column is given its own specific width value
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15


# Function used to restore font details of each column. It takes a specified sheet name and its row limit as input.
def restore_font_details(sheet, row_limit):

    # Restore font details of the header row.
    for row in sheet.iter_rows(max_row=1, max_col=8):
        for cell in row:
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Restore font details of each non-header row.
    for row in sheet.iter_rows(min_row=2, max_row=(row_limit - 1), max_col=8):
        for cell in row:
            cell.font = Font(name='Arial', size=16)


# Function used for applying the cell fill color to appropriate analysis metric columns. It uses a specified sheet
# name and row limit as input.
def apply_column_color(sheet, row_limit):

    # Assign each cell fill color to variables.
    grey_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    light_red_fill = PatternFill(start_color='F08080', end_color='F08080', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # FILL COLOR ASSIGNMENTS FOR THE "Average Completion Time (Start to Finish)" COLUMN
    # Iterate over each row in the specified analysis sheet.
    for row in range(2, row_limit):
 
        # If the row has a value of "N/A", the cell will be filled with grey.
        if sheet['B' + str(row)].value == 'N/A':
            sheet['B' + str(row)].fill = grey_fill

        # If the row has a value that is less than 10, the cell will be filled with green.
        elif sheet['B' + str(row)].value < 10:
            sheet['B' + str(row)].fill = green_fill
        
        # If the row has a value between 10 and 16, the cell will be filled with light green.
        elif 10 <= sheet['B' + str(row)].value <= 16:
            sheet['B' + str(row)].fill = light_green_fill

        # If the row has a value greater than 16 but not greater than 20, the cell will be filled with light red.
        elif 16 < sheet['B' + str(row)].value <= 20:
            sheet['B' + str(row)].fill = light_red_fill
        
        # If the row has a value greater than 20, the cell will be filled with red.
        else:
            sheet['B' + str(row)].fill = red_fill

    # FILL COLOR ASSIGNMENTS FOR THE "Average 811 Call to 811 Mark Time" COLUMN
    # Iterate over each row in the specified analysis sheet.
    for row in range(2, row_limit):

        # If the row has a value of "N/A", the cell will be filled with grey.
        if sheet['C' + str(row)].value == 'N/A':
            sheet['C' + str(row)].fill = grey_fill

        # If the row has a value that is less than 7, the cell will be filled with green.
        elif sheet['C' + str(row)].value < 7:
            sheet['C' + str(row)].fill = green_fill
        
        # If the row has a value between 7 and 14, the cell will be filled with light green.
        elif 7 <= sheet['C' + str(row)].value <= 14:
            sheet['C' + str(row)].fill = light_green_fill

        # If the row has a value greater than 14 but not greater than 20, the cell will be filled with light red.
        elif 14 < sheet['C' + str(row)].value <= 20:
            sheet['C' + str(row)].fill = light_red_fill
        
        # If the row has a value greater than 20, the cell will be filled with red.
        else:
            sheet['C' + str(row)].fill = red_fill

    # FILL COLOR ASSIGNMENTS FOR THE "Average 811 Mark to Completion Time" COLUMN
    # Iterate over each row in the specified analysis sheet.
    for row in range(2, row_limit):

        # If the row has a value of "N/A", the cell will be filled with grey.
        if sheet['D' + str(row)].value == 'N/A':
            sheet['D' + str(row)].fill = grey_fill

        # If the row has a value that is less than 7, the cell will be filled with green.
        elif sheet['D' + str(row)].value < 7:
            sheet['D' + str(row)].fill = green_fill
        
        # If the row has a value between 7 and 14, the cell will be filled with light green.
        elif 7 <= sheet['D' + str(row)].value <= 14:
            sheet['D' + str(row)].fill = light_green_fill

        # If the row has a value greater than 14 but not greater than 20, the cell will be filled with light red.
        elif 14 < sheet['D' + str(row)].value <= 20:
            sheet['D' + str(row)].fill = light_red_fill
        
        # If the row has a value greater than 20, the cell will be filled with red.
        else:
            sheet['D' + str(row)].fill = red_fill

    # FILL COLOR ASSIGNMENTS FOR "% of Jobs Completed on Time (<= 16 Days)" COLUMN
    # Iterate over each row in the specified analysis sheet.
    for row in range(2, row_limit):

        # If the row has a value of "N/A", the cell will be filled with grey.
        if sheet['E' + str(row)].value == 'N/A':
            sheet['E' + str(row)].fill = grey_fill

        # If the row has a value that is greater than or equal to 95, the cell will be filled green.
        elif sheet['E' + str(row)].value >= 95:
            sheet['E' + str(row)].fill = green_fill
        
        # If the row has a value that is at least 90 and less than 95, the cell will be filled light green.
        elif 90 <= sheet['E' + str(row)].value < 95:
            sheet['E' + str(row)].fill = light_green_fill

        # If the row has a value that is at least 80 and less than 90, the cell will be filled light red.
        elif 80 <= sheet['E' + str(row)].value < 90:
            sheet['E' + str(row)].fill = light_red_fill
        
        # If the row has a value that is less than 80, the cell will be filled red.
        else:
            sheet['E' + str(row)].fill = red_fill


# Function used for restoring borders to cells. It uses a specified sheet name and its row limit as input.
def restore_borders(sheet, row_limit):

    # Create the border
    border = Border(
                    top=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    bottom=Side(style='thin'),
                    )
    
    # Iterate over every row in the specifed sheet and apply the border to all cells in every used column.
    for row in sheet.iter_rows(max_row=(row_limit - 1), max_col=8):
        for cell in row:
            cell.border = border


# Function used for restoring the filter tabs in each column. It uses a specified sheet name and its row limit as
# input.
def restore_filters(sheet, row_limit):
    sheet.auto_filter.ref = f"A1:H{row_limit}"

# MAIN FUNCTIONALITY 👇 -----------------------------------------------------------------------------------------#

def restore_analysis_sheets(area_metrics_sheet, month_metrics_sheet):

    # Find the row limit of both the main and archive sheets and assign them to variables.
    area_row_limit = find_row_limit(area_metrics_sheet)
    month_row_limit = find_row_limit(month_metrics_sheet)

    # Restore the row height and column width of both sheets.
    restore_height_and_width(area_metrics_sheet,area_row_limit)
    restore_height_and_width(month_metrics_sheet, month_row_limit)

    # Restore the font details of both sheets.
    restore_font_details(area_metrics_sheet, area_row_limit)
    restore_font_details(month_metrics_sheet, month_row_limit)

    # Apply the cell fill colors of both sheets.
    apply_column_color(area_metrics_sheet, area_row_limit)
    apply_column_color(month_metrics_sheet, month_row_limit)

    # Restore the cell borders of both sheets.
    restore_borders(area_metrics_sheet, area_row_limit)
    restore_borders(month_metrics_sheet, month_row_limit)

    # Restore the filter tabs in both sheets.
    restore_filters(area_metrics_sheet, area_row_limit)
    restore_filters(month_metrics_sheet, month_row_limit)