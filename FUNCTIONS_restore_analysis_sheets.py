
# The purpose of this file is to contain the functions used for the restore_analysis_sheets file.

from openpyxl.styles import Alignment, Font, Border, Side
import FUNCTIONS_cell_color as colors


# FUNCTIONS 👇 ---------------------------------------------------------------------------------------------------#

# This function is used to return the first row of a specified sheet that does not have metric data. It takes a
# sheet name as input.
def find_row_limit(sheet_name):

    # Iterate over the first column column in each row.
    for cell in sheet_name['A']:

        # When the first row that does not have a value in the first column is found, the function will return 
        # that row's number.
        if cell.value is None:
            return(cell.row)
        else:
            continue
    
    # If there are no rows in the specified sheet without an empty first column, return the first next row's number
    return sheet_name.max_row + 1


# Function used for restoring each row's height and each column's width to their original values. It takes a 
# specified sheet name and its row limit as input.
def restore_height_and_width(sheet_name, row_limit):

    # The header row is given a height value of 58
    sheet_name.row_dimensions[1].height = 58

    # Each non-header row is given a height value of 22
    for row in range(2, row_limit):
        sheet_name.row_dimensions[row].height = 22

    # Each column is given its own specific width value
    sheet_name.column_dimensions['A'].width = 15
    sheet_name.column_dimensions['B'].width = 20
    sheet_name.column_dimensions['C'].width = 20
    sheet_name.column_dimensions['D'].width = 20
    sheet_name.column_dimensions['E'].width = 20
    sheet_name.column_dimensions['F'].width = 15
    sheet_name.column_dimensions['G'].width = 15
    sheet_name.column_dimensions['H'].width = 15


# Function used to restore font details of each column. It takes a specified sheet name and its row limit as input.
def restore_font_details(sheet_name, row_limit):

    # Restore font details of the header row.
    for row in sheet_name.iter_rows(max_row=1, max_col=8):
        for cell in row:
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Restore font details of each non-header row.
    for row in sheet_name.iter_rows(min_row=2, max_row=(row_limit - 1), max_col=8):
        for cell in row:
            cell.font = Font(name='Arial', size=16)


# Function used for applying the cell fill color to appropriate analysis metric columns. It uses a specified sheet
# name and row limit as input.
def apply_column_color(sheet_name, row_limit):

    # FILL COLOR ASSIGNMENTS FOR THE "Average Completion Time (Start to Finish)" COLUMN
    # Iterate over each row in the specified analysis sheet.
    for row in range(2, row_limit):
 
        # If the row has a value of "N/A", the cell will be filled with grey.
        if sheet_name['B' + str(row)].value == 'N/A':
            sheet_name['B' + str(row)].fill = colors.grey

        # If the row has a value that is less than 10, the cell will be filled with green.
        elif sheet_name['B' + str(row)].value < 10:
            sheet_name['B' + str(row)].fill = colors.green
        
        # If the row has a value between 10 and 16, the cell will be filled with light green.
        elif 10 <= sheet_name['B' + str(row)].value <= 16:
            sheet_name['B' + str(row)].fill = colors.light_green

        # If the row has a value greater than 16 but not greater than 20, the cell will be filled with light red.
        elif 16 < sheet_name['B' + str(row)].value <= 20:
            sheet_name['B' + str(row)].fill = colors.light_red
        
        # If the row has a value greater than 20, the cell will be filled with red.
        else:
            sheet_name['B' + str(row)].fill = colors.red

    # FILL COLOR ASSIGNMENTS FOR THE "Average 811 Call to 811 Mark Time" COLUMN
    # Iterate over each row in the specified analysis sheet.
    for row in range(2, row_limit):

        # If the row has a value of "N/A", the cell will be filled with grey.
        if sheet_name['C' + str(row)].value == 'N/A':
            sheet_name['C' + str(row)].fill = colors.grey

        # If the row has a value that is less than 3, the cell will be filled with green.
        elif sheet_name['C' + str(row)].value < 3:
            sheet_name['C' + str(row)].fill = colors.green
        
        # If the row has a value between 3 and 5, the cell will be filled with light green.
        elif 3 <= sheet_name['C' + str(row)].value <= 5:
            sheet_name['C' + str(row)].fill = colors.light_green

        # If the row has a value greater than 5 but not greater than 7, the cell will be filled with light red.
        elif 5 < sheet_name['C' + str(row)].value <= 7:
            sheet_name['C' + str(row)].fill = colors.light_red
        
        # If the row has a value greater than 7, the cell will be filled with red.
        else:
            sheet_name['C' + str(row)].fill = colors.red

    # FILL COLOR ASSIGNMENTS FOR THE "Average 811 Mark to Completion Time" COLUMN
    # Iterate over each row in the specified analysis sheet.
    for row in range(2, row_limit):

        # If the row has a value of "N/A", the cell will be filled with grey.
        if sheet_name['D' + str(row)].value == 'N/A':
            sheet_name['D' + str(row)].fill = colors.grey

        # If the row has a value that is less than 4, the cell will be filled with green.
        elif sheet_name['D' + str(row)].value < 4:
            sheet_name['D' + str(row)].fill = colors.green
        
        # If the row has a value between 4 and 7, the cell will be filled with light green.
        elif 4 <= sheet_name['D' + str(row)].value <= 7:
            sheet_name['D' + str(row)].fill = colors.light_green

        # If the row has a value greater than 7 but not greater than 10, the cell will be filled with light red.
        elif 7 < sheet_name['D' + str(row)].value <= 10:
            sheet_name['D' + str(row)].fill = colors.light_red
        
        # If the row has a value greater than 10, the cell will be filled with red.
        else:
            sheet_name['D' + str(row)].fill = colors.red

    # FILL COLOR ASSIGNMENTS FOR "% of Jobs Completed on Time (<= 16 Days)" COLUMN
    # Iterate over each row in the specified analysis sheet.
    for row in range(2, row_limit):

        # If the row has a value of "N/A", the cell will be filled with grey.
        if sheet_name['E' + str(row)].value == 'N/A':
            sheet_name['E' + str(row)].fill = colors.grey

        # If the row has a value that is greater than or equal to 95, the cell will be filled green.
        elif sheet_name['E' + str(row)].value >= 95:
            sheet_name['E' + str(row)].fill = colors.green
        
        # If the row has a value that is at least 90 and less than 95, the cell will be filled light green.
        elif 90 <= sheet_name['E' + str(row)].value < 95:
            sheet_name['E' + str(row)].fill = colors.light_green

        # If the row has a value that is at least 80 and less than 90, the cell will be filled light red.
        elif 80 <= sheet_name['E' + str(row)].value < 90:
            sheet_name['E' + str(row)].fill = colors.light_red
        
        # If the row has a value that is less than 80, the cell will be filled red.
        else:
            sheet_name['E' + str(row)].fill = colors.red


# Function used for restoring borders to cells. It uses a specified sheet name and its row limit as input.
def restore_borders(sheet_name, row_limit):

    # Create the border
    border = Border(
                    top=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    bottom=Side(style='thin'),
                    )
    
    # Iterate over every row in the specifed sheet and apply the border to all cells in every used column.
    for row in sheet_name.iter_rows(max_row=(row_limit - 1), max_col=8):
        for cell in row:
            cell.border = border


# Function used for restoring the filter tabs in each column. It uses a specified sheet name and its row limit as
# input.
def restore_filters(sheet, row_limit):
    sheet.auto_filter.ref = f"A1:H{row_limit}"
