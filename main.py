
# This is an automation script I wrote during my time as an intern at Knoxville Utilities Board's fiber 
# department. What this script does is automates the update of the four following sheets in the Fiber Installations
# Database Excel Spreadsheet:
#   - Main installation job data sheet (currently titled "Main Installs")
#   - Installation job archive sheet of jobs completed for at least 90 days (currently titled ">90 Day Archive")
#   - Analysis sheet used to house specific job metrics for each work area present in the spreadsheet (currently 
#     titled "Area Metrics")
#   - Analysis sheet used to house specific job metrics for each month present in the spreadsheet (currently titled 
#     "Month-by-Month Metrics")
# In the workplace, this file is automated by using an API to connect to KUB's azure account to retrieve, update,
# and reupload the file on a daily basis. However, this version just requires this file to be ran.
# NOTE: The data in this Excel file is NOT actual company data as it has been replaced with sample data, but the
#       functionality is identical to the currently in-use version.

import pandas as pd
from openpyxl import load_workbook

import warnings
from time import sleep, time

from check_backup_directory_and_run_time_log import check_backup_directory_and_run_time_log
from update_main_and_archive import update_main_and_archive
from update_area_metrics import update_area_metrics
from update_month_metrics import update_month_metrics
from restore_main_and_archive import restore_main_and_archive
from restore_analysis_sheets import restore_analysis_sheets


# Step 1. Save a backup for the current iteration👇 --------------------------------------------------------------#

start_time = time()

# Verify that the current month's backup directory and run time log exist.
print("\nChecking if backup directory exists...")
current_date, month_key = check_backup_directory_and_run_time_log()
print(f"\nBackup directory for {month_key} verified!")
sleep(1)

# Load the Excel file in OpenpyXL
print("\nLoading Excel file in OpenpyXL...")
workbook = load_workbook("Fiber Installations Database - Pre Update.xlsx")
print("\nExcel file loaded in OpenpyXL!")
sleep(1)

# Save a backup of the Excel file for the current date.
print(f"\nSaving backup of Excel file for date: {current_date}...")
workbook.save(f"Backups\\{month_key}\\Backup - {current_date}.xlsx")
print(f"\nBackup for {current_date} created!")
sleep(1)


# Step 2. Modify sheet data with Pandas 👇 -----------------------------------------------------------------------#

warnings.filterwarnings('ignore', category=FutureWarning)

# Load the main and archive sheets as Pandas dataframes
print("\nLoading Excel sheets in Pandas...")
df_main = pd.read_excel("Fiber Installations Database - Pre Update.xlsx", sheet_name='Main Installs')
df_90day = pd.read_excel("Fiber Installations Database - Pre Update.xlsx", sheet_name='>90 Day Archive')
print("\nExcel sheets loaded in Pandas!")
sleep(1)

# Update the data in the main and archive sheets
print("\nUpdating Main and Archive Sheet...")
updated_main, updated_90day = update_main_and_archive(df_main, df_90day)
print("\nMain and Archive Sheet updated!")
sleep(1)

# Update the data in the area metrics analysis sheet.
print("\nUpdating Area Metrics sheet...")
updated_area_metrics = update_area_metrics(updated_main, updated_90day)
print("\nArea Metrics sheet updated!")
sleep(1)

# Update the data in the month-by-month metrics analysis sheet.
print("\nUpdating Month-by-Month Metrics sheet...")
updated_month_metrics = update_month_metrics(updated_main, updated_90day)
print("\nMonth-by-Month Metrics sheet updated!")
sleep(1)

# Save changes made by Pandas functionality to Excel file.
print("\nSaving Pandas modifications...")
with pd.ExcelWriter("Fiber Installations Database - Post Update.xlsx",
                    engine='openpyxl',
                    mode='a',
                    if_sheet_exists='replace') as writer:
        
    updated_main.to_excel(writer, sheet_name='Main Installs', index=False)
    updated_90day.to_excel(writer, sheet_name='>90 Day Archive', index=False)
    updated_area_metrics.to_excel(writer, sheet_name='Area Metrics', index=False)
    updated_month_metrics.to_excel(writer, sheet_name='Month-by-Month Metrics', index=False)
print("\nPandas modifications saved!")
sleep(1)


# Step 3. Restore sheet Formatting with OpenpyXL 👇 --------------------------------------------------------------#

# Load all four modified sheets in OpenpyXL.
print("\nLoading Excel sheets in OpenpyXL...")
workbook = load_workbook("Fiber Installations Database - Post Update.xlsx")
main_sheet = workbook['Main Installs']
archive_sheet = workbook['>90 Day Archive']
area_metrics_sheet = workbook['Area Metrics']
month_metrics_sheet = workbook['Month-by-Month Metrics']
print("\nExcel sheets loaded in OpenpyXL!")
sleep(1)

# Restore the main and archive sheet formating.
print("\nRestoring main and archive sheet formatting...")
restore_main_and_archive(main_sheet, archive_sheet)
print("\nMain and archive sheet formatting restored!")
sleep(1)

# Restore the area metrics and month-by-month metrics analysis sheet formatting.
print("\nRestoring area and month-by-month metrics analysis sheet formatting...")
restore_analysis_sheets(area_metrics_sheet, month_metrics_sheet)
print("\nArea and month-by-month metrics analysis sheet formatting restored!")
sleep(1)

# Save changes made by OpenpyXL functionality to Excel file.
print("\nSaving OpenpyXL modifications...")
workbook.save("Fiber Installations Database - Post Update.xlsx")
print("\nOpenpyXL modifications saved!")

# Display the run time of the current iteration of the automation process.
end_time = time()
run_time = round((end_time - start_time), 2)
print(f"Run time: {run_time} seconds.")

# Append the run time of the current date's automation process into the current month's run time log.
with open(f'Run Times\\{month_key}.txt', 'a') as file:
    file.write(f"\n{current_date}: {run_time} seconds")
