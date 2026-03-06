
# The purpose of this file is to contain the functions used for the restore_main_and_archive file.

from openpyxl.styles import Alignment, Font, Border, Side
import FUNCTIONS_cell_color as colors
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule
import re

# FUNCTIONS 👇 ---------------------------------------------------------------------------------------------------#


# Function used for restoring each row's height and each column's width to their original values. It takes a 
# specified sheet name as input.
def restore_height_and_width(sheet_name):

    # The header row is given a height value of 35
    sheet_name.row_dimensions[1].height = 35

    # Each non-header row is set to autofit the row depending on if the row has wrapped text or not.
    for row in range(2, sheet_name.max_row + 1):
        sheet_name.row_dimensions[row].height = None

    # Each column is given its own specific width value
    sheet_name.column_dimensions['A'].width = 17.5
    sheet_name.column_dimensions['B'].width = 17.5
    sheet_name.column_dimensions['C'].width = 17.5
    sheet_name.column_dimensions['D'].width = 17.5
    sheet_name.column_dimensions['E'].width = 17.5
    sheet_name.column_dimensions['F'].width = 17.5
    sheet_name.column_dimensions['G'].width = 27
    sheet_name.column_dimensions['H'].width = 60
    sheet_name.column_dimensions['I'].width = 15
    sheet_name.column_dimensions['J'].width = 25
    sheet_name.column_dimensions['K'].width = 25
    sheet_name.column_dimensions['L'].width = 40
    sheet_name.column_dimensions['M'].width = 50
    sheet_name.column_dimensions['N'].width = 70
    sheet_name.column_dimensions['O'].width = 35
    sheet_name.column_dimensions['P'].width = 9
    sheet_name.column_dimensions['Q'].width = 25


    # Each of the date containing rows has text wrapping enabled in order to maximize visibility.
    for row in sheet_name.iter_rows(min_row=1, max_col=1, min_col=6):
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical,
                wrap_text=True
            )

    # Each of the date containing rows has text wrapping enabled in order to maximize visibility.
    for row in sheet_name.iter_rows(min_row=1, max_col=14, min_col=14):
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical,
                wrap_text=True
            )


# Function used to restore font details of each column. It takes a specified sheet name as input.
def restore_font_details(sheet_name):

    # Restore font details of the header row.
    for row in sheet_name.iter_rows(max_row=1, max_col=15):
        for cell in row:
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Restore each date value to mm-dd-yyyy format.
    for row in sheet_name.iter_rows(min_row=2, max_row=sheet_name.max_row, max_col=7):
        for cell in row:
            cell.number_format = 'mm-dd-yyyy'

    # Restore font details of columns A-M
    for row in sheet_name.iter_rows(min_row=2, max_row=sheet_name.max_row, max_col=13):
        for cell in row:
            cell.font = Font(name='Arial', size=16)

    # Restore font details of the "Notes" column.
    for row in sheet_name.iter_rows(min_row=2, max_row=sheet_name.max_row, min_col=14, max_col=14):
        for cell in row:
            cell.font = Font(name='Arial', size=11, bold=True)

    # Restore font details of the "Status" Column.
    for row in sheet_name.iter_rows(min_row=2, max_row=sheet_name.max_row, min_col=15, max_col=15):
        for cell in row:
            if cell.value is not None:
                cell.value = str(cell.value).upper()
            cell.font = Font(name='Arial', size=16, bold=True)


# Function used for restoring cell phone number formats in the "Contact #" column. It takes a sheet name as input.
def restore_phone_numbers(sheet_name):

    # Iterate over each non-header row and get the contact phone number value for each row.
    for row in range(2, sheet_name.max_row + 1):
        cell = sheet_name[f"K{row}"]

        # The the "Contact #" column has a value, convert the value into a numeric format.
        if cell.value:
            
            # Remove all non numeric entries.
            digits = re.sub(r"\D", "", str(cell.value))

            # If the length of the stripped value is 10 digits, convert the value into an integer.
            if len(digits) == 10:
                cell.value = int(digits)
        
        # Format the value as a cell phone number format.
        cell.number_format = "(###) ###-####"


# Function used for restoring the cell fill color to each cell in the "Job Completed Date" and "Address" columns. It
# takes a sheet name as input.
def restore_column_color(sheet_name):

    # Iterate over the 'Status' column in each row of the specified sheet.
    for row in range(2, sheet_name.max_row + 1):

        # If the row has a status of "N/A", then the cell will have no fill.
        if sheet_name['O' + str(row)].value is None:
            sheet_name['G' + str(row)].fill = colors.white
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.white
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.white
            sheet_name['O' + str(row)].font = Font(color='000000')

        # If the row has a status of "COMPLETED", the two cells will be filled with blue.
        elif sheet_name['O' + str(row)].value.upper() == 'COMPLETED':
            sheet_name['G' + str(row)].fill = colors.blue
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.blue
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.blue
            sheet_name['O' + str(row)].font = Font(color='000000')

        # If the row has a status of "HIGH PRIORITY", the two cells will be filled with magenta.
        elif sheet_name['O' + str(row)].value.upper() == 'HIGH PRIORITY':
            sheet_name['G' + str(row)].fill = colors.magenta
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.magenta
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.magenta
            sheet_name['O' + str(row)].font = Font(color='000000')

        # If the row has a status of "KUB/GLOBAL", the two cells will be filled with cyan.
        elif sheet_name['O' + str(row)].value.upper() == 'KUB/GLOBAL':
            sheet_name['G' + str(row)].fill = colors.cyan
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.cyan
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.cyan
            sheet_name['O' + str(row)].font = Font(color='000000')

        # If the row has a status of "NEED TO CALL 811", the two cells will be filled with plum.
        elif sheet_name['O' + str(row)].value.upper() == 'NEED TO CALL 811':
            sheet_name['G' + str(row)].fill = colors.plum
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.plum
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.plum
            sheet_name['O' + str(row)].font = Font(color='000000')

        # If the row has a status of "WAITING ON 811", the two cells will be filled with yellow.
        elif sheet_name['O' + str(row)].value.upper() == 'WAITING ON 811':
            sheet_name['G' + str(row)].fill = colors.yellow
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.yellow
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.yellow
            sheet_name['O' + str(row)].font = Font(color='000000')

        # If the row has a status of "WAITING ON 811", the two cells will be filled with yellow.
        elif sheet_name['O' + str(row)].value.upper() == 'WAITING ON PAPERWORK':
            sheet_name['G' + str(row)].fill = colors.silver
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.silver
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.silver
            sheet_name['O' + str(row)].font = Font(color='000000')

        # If the row has a status of "NOTES", the two cells will be filled with gold.
        elif sheet_name['O' + str(row)].value.upper() == 'NOTES':
            sheet_name['G' + str(row)].fill = colors.ut_orange
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.ut_orange
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.ut_orange
            sheet_name['O' + str(row)].font = Font(color='000000')

        # If the row has a status of "ON HOLD/WAITING ON CUST TO CALL", the two cells will be filled with olive.
        elif sheet_name['O' + str(row)].value.upper() == 'ON HOLD/WAITING ON CUST TO CALL':
            sheet_name['G' + str(row)].fill = colors.brown
            sheet_name['G' + str(row)].font = Font(color='ffffff')
            sheet_name['H' + str(row)].fill = colors.brown
            sheet_name['H' + str(row)].font = Font(color='ffffff')
            sheet_name['O' + str(row)].fill = colors.brown
            sheet_name['O' + str(row)].font = Font(color='ffffff')

        # If the row has a status of "READY TO BURY", the two cells will be filled with yellow green.
        elif sheet_name['O' + str(row)].value.upper() == 'READY TO BURY':
            sheet_name['G' + str(row)].fill = colors.lime_green
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.lime_green
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.lime_green
            sheet_name['O' + str(row)].font = Font(color='000000')

        # If the row has a status of "SCHEDULED", the two cells will be filled with red.
        elif sheet_name['O' + str(row)].value.upper() == 'SCHEDULED':
            sheet_name['G' + str(row)].fill = colors.red
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.red
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.red
            sheet_name['O' + str(row)].font = Font(color='000000')
        
        # If the row has a status of "NEXT DAY", the two cells will be filled with red.
        elif sheet_name['O' + str(row)].value.upper() == 'NEXT DAY':
            sheet_name['G' + str(row)].fill = colors.coral
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.coral
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.coral
            sheet_name['O' + str(row)].font = Font(color='000000')

        # If the row has a status of "CANCELLED", the two cells will be filled with black and the text font color
        # will be changed to white and have a strikethrough effect.
        elif sheet_name['O' + str(row)].value.upper() == 'CANCELLED':
            sheet_name['G' + str(row)].fill = colors.grey
            sheet_name['G' + str(row)].font = Font(color='ffffff')
            sheet_name['H' + str(row)].fill = colors.grey
            sheet_name['H' + str(row)].font = Font(color='ffffff')
            sheet_name['O' + str(row)].fill = colors.grey
            sheet_name['O' + str(row)].font = Font(color='ffffff')
        
        # If the Status column value for the row is something else other than the previous listed options,
        # then the cell will have no fill
        else:
            sheet_name['G' + str(row)].fill = colors.white
            sheet_name['G' + str(row)].font = Font(color='000000')
            sheet_name['H' + str(row)].fill = colors.white
            sheet_name['H' + str(row)].font = Font(color='000000')
            sheet_name['O' + str(row)].fill = colors.white
            sheet_name['O' + str(row)].font = Font(color='000000')

        # If the "Notes" column has a value, then the cell is filled with gold.
        if sheet_name['N' + str(row)].value is not None:
            sheet_name['N' + str(row)].fill = colors.ut_orange


# Function used for restoring borders to cells. It uses a specified sheet name as input.
def restore_borders(sheet_name):

    # Create the border.
    border = Border(
                    top=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    bottom=Side(style='thin'),
                    )
    
    # Iterate over every row in the specifed sheet and apply the border to all cells in every used column.
    for row in sheet_name.iter_rows(max_row=sheet_name.max_row, max_col=15):
        for cell in row:
            cell.border = border


# Function used for restoring the filter tabs in each column. It uses a specified sheet name as input.
def restore_filters(sheet_name):
    sheet_name.auto_filter.ref = f"A1:O{sheet_name.max_row}"


# Function used for restoring the data validation tabs of the 'Status' column. It uses a specified sheet name, row limit,
# and openpyxl workbook as input.
def restore_data_validation(sheet_name, workbook):

    # Create a list with all job statuses to be used later.
    status_list = [
    "N/A",
    "COMPLETED",
    "HIGH PRIORITY",
    "KUB/GLOBAL",
    "NEED TO CALL 811",
    "WAITING ON 811",
    "WAITING ON PAPERWORK",
    "NOTES",
    "ON HOLD/WAITING ON CUST TO CALL",
    "READY TO BURY",
    "SCHEDULED",
    "NEXT DAY",
    "CANCELLED"
    ]
    
    LIST_SHEET = "_validation_lists"
    LIST_NAME = "StatusList"

    # Access a hidden sheet saved under tha variable, LIST_SHEET
    if LIST_SHEET in workbook.sheetnames:
        ls = workbook[LIST_SHEET]
    else:
        ls = workbook.create_sheet(LIST_SHEET)
        ls.sheet_state = "hidden"

    # Clear all statuses saved in the list sheet and recreate them. This allows us
    # to ensure the statuses that are available in the data validation functionality
    # are up to date.
    clear_to = max(ls.max_row, len(status_list)) + 5
    for r in range(1, clear_to + 1):
        ls[f"A{r}"].value = None
    for i, status in enumerate(status_list, start=1):
        ls[f"A{i}"].value = status

    # Force-refresh the defined name to the full list range
    if LIST_NAME in workbook.defined_names:
        del workbook.defined_names[LIST_NAME]

    # define named range
    ref = f"'{LIST_SHEET}'!$A$1:$A${len(status_list)}"
    workbook.defined_names.add(DefinedName(LIST_NAME, attr_text=ref))

    # Create a DataValidation class instance to be used.
    dv = DataValidation(
        type='list',
        formula1=f"={LIST_NAME}",
        allow_blank=True
    )

    # Add the data validation to the 'Status' column
    sheet_name.add_data_validation(dv)
    dv.add(f'O2:O{sheet_name.max_row}')


# Function used to restore conditional formatting to status column. It uses a specified
# sheet name and its row limit as inputs.
def restore_conditional_formatting(sheet_name):

    # Remove the formatting rules from previous runs to ensure no errors due to formatting
    # merging.
    sheet_name.conditional_formatting._cf_rules.clear()

    # The range of cells that will be affected by the conditional formatting.
    rng = f"G2:H{sheet_name.max_row} O2:O{sheet_name.max_row}"

    # Establish rules for each status option in the 'Status' column.
    rules = [
        ("N/A", colors.white, "000000"),
        ("COMPLETED", colors.blue, "000000"),
        ("HIGH PRIORITY", colors.magenta, "000000"),
        ("KUB/GLOBAL", colors.cyan, "000000"),
        ("NEED TO CALL 811", colors.plum, "000000"),
        ("WAITING ON 811", colors.yellow, "000000"),
        ("WAITING ON PAPERWORK", colors.silver, "000000"),
        ("NOTES", colors.ut_orange, "000000"),
        ("ON HOLD/WAITING ON CUST TO CALL", colors.brown, "FFFFFF"),
        ("READY TO BURY", colors.lime_green, "000000"),
        ("SCHEDULED", colors.red, "000000"),
        ("NEXT DAY", colors.coral, "000000"),
        ("CANCELLED", colors.grey, "FFFFFF")
    ]

    # For each status in the rule list, create the Excel formula for them.
    for status_text, fill, font_color in rules:

        rule = FormulaRule(
            formula=[f'=$O2="{status_text}"'],
            fill=fill,
            font=Font(color=font_color),
            stopIfTrue=True
        )

        # Add the conditional formatting rules to the sheet.
        sheet_name.conditional_formatting.add(rng, rule)


# Function used for restoring the Legend of the main sheet. It takes the main sheet name as input.
def restore_legend(main_sheet):

    # Restore the legend colors
    main_sheet['P2'].fill = colors.blue
    main_sheet['P3'].fill = colors.magenta
    main_sheet['P4'].fill = colors.cyan
    main_sheet['P5'].fill = colors.plum
    main_sheet['P6'].fill = colors.yellow
    main_sheet['P7'].fill = colors.silver
    main_sheet['P8'].fill = colors.ut_orange
    main_sheet['P9'].fill = colors.brown
    main_sheet['P10'].fill = colors.lime_green
    main_sheet['P11'].fill = colors.red
    main_sheet['P12'].fill = colors.coral
    main_sheet['P13'].fill = colors.grey

    # Restore the legend labels
    main_sheet['Q2'].value = "Completed"
    main_sheet['Q3'].value = "High Priority"
    main_sheet['Q4'].value = "KUB / Global"
    main_sheet['Q5'].value = "Need to Call 811"
    main_sheet['Q6'].value = "Waiting on 811"
    main_sheet['Q7'].value = 'Waiting on Paperwork'
    main_sheet['Q8'].value = "Notes"
    main_sheet['Q9'].value = "On Hold / Waiting on Customer to Call"
    main_sheet['Q10'].value = "Ready to Bury"
    main_sheet['Q11'].value = "Scheduled"
    main_sheet['Q12'].value = "Next Day"
    main_sheet['Q13'].value = "Cancelled"

    # Restore font details for the legend labels.
    for row in main_sheet.iter_rows(min_row=3, max_row=9, min_col=17, max_col=17):
        for cell in row:
            cell.font = Font(name='Calibri', size=12)


# Function used for restoring the filter tabs in each column. It uses a specified sheet name as input.
def restore_filters(sheet_name):
    sheet_name.auto_filter.ref = f"A1:O{sheet_name.max_row}"
    sheet_name.freeze_panes = "A2"
