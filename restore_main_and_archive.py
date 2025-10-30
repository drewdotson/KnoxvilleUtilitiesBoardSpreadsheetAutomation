
# The purpose of this file is to restore the pre-update format of the main and archive sheets.

from openpyxl.styles import Alignment, PatternFill, Font, Border, Side


# FUNCTIONS 👇 ---------------------------------------------------------------------------------------------------#

# This function is used to return the first row of a specified sheet that does not have installation job data.
# It takes a sheet name as input.
def find_row_limit(sheet):

    # Iterate over the "Status" column in every row.
    for cell in sheet['N']:

        # When the first row that does not have a value in the "Status" column is found, the function will return 
        # that row's number.
        if cell.value is None:
            return(cell.row)
        else:
            continue
    
    # If there are no rows in the specified sheet without an empty "Status" column, return the first empty row's 
    # number
    return sheet.max_row + 1


# Function used for restoring each row's height and each column's width to their original values. It takes a 
# specified sheet name and its row limit as input.
def restore_height_and_width(sheet, row_limit):

    # The header row is given a height value of 35
    sheet.row_dimensions[1].height = 35

    # Each non-header row is given a height value of 30
    for row in range(2, row_limit):
        sheet.row_dimensions[row].height = 30

    # Each column is given its own specific width value
    sheet.column_dimensions['A'].width = 32
    sheet.column_dimensions['B'].width = 34
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 32
    sheet.column_dimensions['E'].width = 40
    sheet.column_dimensions['F'].width = 37
    sheet.column_dimensions['G'].width = 60
    sheet.column_dimensions['H'].width = 15
    sheet.column_dimensions['I'].width = 25
    sheet.column_dimensions['J'].width = 25
    sheet.column_dimensions['K'].width = 40
    sheet.column_dimensions['L'].width = 50
    sheet.column_dimensions['M'].width = 70
    sheet.column_dimensions['N'].width = 35
    sheet.column_dimensions['O'].width = 9
    sheet.column_dimensions['P'].width = 25


# Function used to restore font details of each column. It takes a specified sheet name and its row limit as input.
def restore_font_details(sheet, row_limit):

    # Restore font details of the header row.
    for row in sheet.iter_rows(max_row=1, max_col=14):
        for cell in row:
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Restore each date value to mm-dd-yyyy format.
    for row in sheet.iter_rows(min_row=2, max_row=(row_limit - 1), max_col=6):
        for cell in row:
            cell.number_format = 'mm-dd-yyyy'

    # Restore font details of columns A-L
    for row in sheet.iter_rows(min_row=2, max_row=(row_limit - 1), max_col=12):
        for cell in row:
            cell.font = Font(name='Arial', size=16)

    # Restore font details of the "Notes" column.
    for row in sheet.iter_rows(min_row=2, max_row=(row_limit - 1), min_col=13, max_col=13):
        for cell in row:
            cell.font = Font(name='Arial', size=11, bold=True)

    # Restore font details of the "Status" Column.
    for row in sheet.iter_rows(min_row=2, max_row=(row_limit - 1), min_col=14, max_col=14):
        for cell in row:
            cell.font = Font(name='Arial', size=16, bold=True)


# Function used for restoring the cell fill color to each cell in the "Job Completed Date" and "Address" columns. It
# takes a sheet name and its row limit as input.
def restore_column_color(sheet, row_limit):

    # Assign each cell fill color to variables.
    blue_fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
    magenta_fill = PatternFill(start_color='FF00FF', end_color='FF00FF', fill_type='solid')
    cyan_fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    plum_fill = PatternFill(start_color='DDA0DD', end_color='DDA0DD', fill_type='solid')
    yellow_fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
    gold_fill = PatternFill(start_color='ffd700', end_color='ffd700', fill_type='solid')
    olive_fill = PatternFill(start_color='808000', end_color='808000', fill_type='solid')
    yellow_green_fill = PatternFill(start_color='9ACD32', end_color='9ACD32', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    # Iterate over the 'Status' column in each row of the specified sheet.
    for row in range(2, row_limit):

        # If the row has a status of "COMPLETED", the two cells will be filled with blue.
        if sheet['N' + str(row)].value == 'COMPLETED':
            sheet['F' + str(row)].fill = blue_fill
            sheet['G' + str(row)].fill = blue_fill
        
        # If the row has a status of "HIGH PRIORITY", the two cells will be filled with magenta.
        elif sheet['N' + str(row)].value == 'HIGH PRIORITY':
            sheet['F' + str(row)].fill = magenta_fill
            sheet['G' + str(row)].fill = magenta_fill

        # If the row has a status of "KUB/GLOBAL", the two cells will be filled with cyan.
        elif sheet['N' + str(row)].value == 'KUB/GLOBAL':
            sheet['F' + str(row)].fill = cyan_fill
            sheet['G' + str(row)].fill = cyan_fill

        # If the row has a status of "NEED TO CALL 811", the two cells will be filled with plum.
        elif sheet['N' + str(row)].value == 'NEED TO CALL 811':
            sheet['F' + str(row)].fill = plum_fill
            sheet['G' + str(row)].fill = plum_fill

        # If the row has a status of "WAITING ON 811", the two cells will be filled with yellow.
        elif sheet['N' + str(row)].value == 'WAITING ON 811':
            sheet['F' + str(row)].fill = yellow_fill
            sheet['G' + str(row)].fill = yellow_fill

        # If the row has a status of "NOTES", the two cells will be filled with gold.
        elif sheet['N' + str(row)].value == 'NOTES':
            sheet['F' + str(row)].fill = gold_fill
            sheet['G' + str(row)].fill = gold_fill

        # If the row has a status of "ON HOLD/WAITING ON CUST TO CALL", the two cells will be filled with olive.
        elif sheet['N' + str(row)].value == 'ON HOLD/WAITING ON CUST TO CALL':
            sheet['F' + str(row)].fill = olive_fill
            sheet['G' + str(row)].fill = olive_fill

        # If the row has a status of "READY TO BURY", the two cells will be filled with yellow green.
        elif sheet['N' + str(row)].value == 'READY TO BURY':
            sheet['F' + str(row)].fill = yellow_green_fill
            sheet['G' + str(row)].fill = yellow_green_fill

        # If the row has a status of "SCHEDULED", the two cells will be filled with red.
        elif sheet['N' + str(row)].value == 'SCHEDULED':
            sheet['F' + str(row)].fill = red_fill
            sheet['G' + str(row)].fill = red_fill

        # If the row has a status of "CANCELLED", the two cells will be filled with black and the text font color
        # will be changed to white and have a strikethrough effect.
        elif sheet['N' + str(row)].value == 'CANCELLED':
            sheet['F' + str(row)].fill = black_fill
            sheet['F' + str(row)].font = Font(color='ffffff', strike=True)
            sheet['G' + str(row)].fill = black_fill
            sheet['G' + str(row)].font = Font(color='ffffff', strike=True)

        # If the "Notes" column has a value, then the cell is filled with gold.
        if sheet['M' + str(row)].value is not None:
            sheet['M' + str(row)].fill = gold_fill


# Function used for restoring borders to cells. It uses a specified sheet name and its row limit as input.
def restore_borders(sheet, row_limit):

    # Create the border.
    border = Border(
                    top=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    bottom=Side(style='thin'),
                    )
    
    # Iterate over every row in the specifed sheet and apply the border to all cells in every used column.
    for row in sheet.iter_rows(max_row=(row_limit - 1), max_col=14):
        for cell in row:
            cell.border = border


# Function used for restoring the filter tabs in each column. It uses a specified sheet name and its row limit as
# input.
def restore_filters(sheet, row_limit):
    sheet.auto_filter.ref = f"A1:N{row_limit}"


# Function used for restoring the Legend of the main sheet. It takes the main sheet name as input.
def restore_legend(main_sheet):

    # Assign each cell fill color to variables.
    blue_fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
    magenta_fill = PatternFill(start_color='FF00FF', end_color='FF00FF', fill_type='solid')
    cyan_fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    plum_fill = PatternFill(start_color='DDA0DD', end_color='DDA0DD', fill_type='solid')
    yellow_fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
    gold_fill = PatternFill(start_color='ffd700', end_color='ffd700', fill_type='solid')
    olive_fill = PatternFill(start_color='808000', end_color='808000', fill_type='solid')
    yellow_green_fill = PatternFill(start_color='9ACD32', end_color='9ACD32', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    # Restore the legend colors
    main_sheet['O3'].fill = blue_fill
    main_sheet['O4'].fill = magenta_fill
    main_sheet['O5'].fill = cyan_fill
    main_sheet['O6'].fill = plum_fill
    main_sheet['O7'].fill = yellow_fill
    main_sheet['O8'].fill = gold_fill
    main_sheet['O9'].fill = olive_fill
    main_sheet['O10'].fill = yellow_green_fill
    main_sheet['O11'].fill = red_fill
    main_sheet['O12'].fill = black_fill

    # Restore the legend labels
    main_sheet['P3'].value = "Completed"
    main_sheet['P4'].value = "High Priority"
    main_sheet['P5'].value = "KUB / Global"
    main_sheet['P6'].value = "Need to Call 811"
    main_sheet['P7'].value = "Waiting on 811"
    main_sheet['P8'].value = "Notes"
    main_sheet['P9'].value = "On Hold / Waiting on Customer to Call"
    main_sheet['P10'].value = "Ready to Bury"
    main_sheet['P11'].value = "Scheduled"
    main_sheet['P12'].value = "Cancelled"

    # Restore font details for the legend labels.
    for row in main_sheet.iter_rows(min_row=3, max_row=9, min_col=16, max_col=16):
        for cell in row:
            cell.font = Font(name='Calibri', size=12)


# Function used for restoring the filter tabs in each column. It uses a specified sheet name and its row limit as
# input.
def restore_filters(sheet, row_limit):
    sheet.auto_filter.ref = f"A1:N{row_limit}"


# MAIN FUNCTIONALITY 👇 -----------------------------------------------------------------------------------------#

def restore_main_and_archive(main_sheet, archive_sheet):

    # Find the row limit of both the main and archive sheets and assign them to variables.
    main_row_limit = find_row_limit(main_sheet)
    archive_row_limit = find_row_limit(archive_sheet)

    # Restore the row height and column width of both sheets.
    restore_height_and_width(main_sheet,main_row_limit)
    restore_height_and_width(archive_sheet, archive_row_limit)

    # Restore the font details of both sheets.
    restore_font_details(main_sheet, main_row_limit)
    restore_font_details(archive_sheet, archive_row_limit)

    # Restore the cell fill colors of both sheets.
    restore_column_color(main_sheet, main_row_limit)
    restore_column_color(archive_sheet, archive_row_limit)

    # Restore the cell borders of both sheets.
    restore_borders(main_sheet, main_row_limit)
    restore_borders(archive_sheet, archive_row_limit)

    # Restore the cell borders of both sheets.
    restore_borders(main_sheet, main_row_limit)
    restore_borders(archive_sheet, archive_row_limit)

    # Restore the filter tabs in both sheets.
    restore_filters(main_sheet, main_row_limit)
    restore_filters(archive_sheet, archive_row_limit)

    # Restore the cell fill legend of the main sheet.
    restore_legend(main_sheet)
